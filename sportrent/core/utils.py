"""
Утилиты для экспорта данных в XLSX и PDF.
"""

import os
import sys
from io import BytesIO
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Шрифт с поддержкой кириллицы для PDF
_PDF_CYRILLIC_FONT = None


def _get_pdf_cyrillic_font():
    """
    Регистрирует и возвращает шрифт с поддержкой кириллицы для PDF.
    Порядок: 1) Встроенные шрифты в static/fonts (одинаково на Windows и Linux),
    2) На Windows — Arial из системы, 3) Helvetica (кириллица не отображается).
    Для работы на Linux поместите DejaVuSans.ttf и DejaVuSans-Bold.ttf в static/fonts/.
    """
    global _PDF_CYRILLIC_FONT
    if _PDF_CYRILLIC_FONT:
        return _PDF_CYRILLIC_FONT
    font_name = 'Helvetica'
    try:
        base_dir = getattr(settings, 'BASE_DIR', None)
        if base_dir:
            base_dir = str(base_dir)
            # Единый путь для всех ОС: проект/static/fonts (рядом с manage.py — корень проекта)
            fonts_dir = os.path.join(base_dir, 'static', 'fonts')
            normal_path = os.path.join(fonts_dir, 'DejaVuSans.ttf')
            bold_path = os.path.join(fonts_dir, 'DejaVuSans-Bold.ttf')
            if os.path.exists(normal_path):
                pdfmetrics.registerFont(TTFont('PdfCyrillic', normal_path))
                pdfmetrics.registerFont(TTFont('PdfCyrillicBold', bold_path if os.path.exists(bold_path) else normal_path))
                _PDF_CYRILLIC_FONT = 'PdfCyrillic'
                return _PDF_CYRILLIC_FONT
        if sys.platform == 'win32':
            win_fonts = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
            arial_path = os.path.join(win_fonts, 'arial.ttf')
            arial_bold_path = os.path.join(win_fonts, 'arialbd.ttf')
            if os.path.exists(arial_path):
                pdfmetrics.registerFont(TTFont('PdfCyrillic', arial_path))
                pdfmetrics.registerFont(TTFont('PdfCyrillicBold', arial_bold_path if os.path.exists(arial_bold_path) else arial_path))
                _PDF_CYRILLIC_FONT = 'PdfCyrillic'
                return _PDF_CYRILLIC_FONT
        # Linux: типичные пути к DejaVu (пакет fonts-dejavu-core)
        for linux_dir in ('/usr/share/fonts/truetype/dejavu', '/usr/share/fonts/TTF', '/usr/share/fonts/dejavu'):
            normal_path = os.path.join(linux_dir, 'DejaVuSans.ttf')
            bold_path = os.path.join(linux_dir, 'DejaVuSans-Bold.ttf')
            if os.path.exists(normal_path):
                pdfmetrics.registerFont(TTFont('PdfCyrillic', normal_path))
                pdfmetrics.registerFont(TTFont('PdfCyrillicBold', bold_path if os.path.exists(bold_path) else normal_path))
                _PDF_CYRILLIC_FONT = 'PdfCyrillic'
                return _PDF_CYRILLIC_FONT
    except Exception:
        pass
    _PDF_CYRILLIC_FONT = font_name
    return _PDF_CYRILLIC_FONT


def export_inventory_to_xlsx(inventory_qs):
    """
    Экспорт списка инвентаря в XLSX.
    """
    # Создаем workbook и активный лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Инвентарь"

    # Заголовки
    headers = ['Название', 'Категория', 'Бренд', 'Модель', 'Состояние', 'Цена/день', 'Статус', 'Рейтинг', 'Аренд']
    ws.append(headers)

    # Стили заголовка
    header_fill = PatternFill(start_color='2C5F7C', end_color='2C5F7C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Данные
    for item in inventory_qs:
        ws.append([
            item.name,
            item.category.name,
            item.brand or '-',
            item.model or '-',
            item.get_condition_display(),
            float(item.price_per_day),
            item.get_status_display(),
            float(item.avg_rating) if item.avg_rating else '-',
            item.total_rentals,
        ])

    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Создаем HTTP response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventory_{datetime.now().strftime("%Y%m%d")}.xlsx"'

    return response


def export_rentals_to_xlsx(rentals_qs):
    """
    Экспорт списка аренд в XLSX.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Аренды"

    # Заголовки
    headers = ['ID', 'Инвентарь', 'Клиент', 'Дата начала', 'Дата окончания', 'Дней', 'Стоимость', 'Статус']
    ws.append(headers)

    # Стили
    header_fill = PatternFill(start_color='3A7CA5', end_color='3A7CA5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Данные
    for rental in rentals_qs:
        ws.append([
            str(rental.rental_id)[:8],
            rental.inventory.name,
            rental.client.full_name,
            rental.start_date.strftime('%d.%m.%Y'),
            rental.end_date.strftime('%d.%m.%Y'),
            rental.rental_days,
            float(rental.total_price),
            rental.get_status_display(),
        ])

    # Автоширина
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rentals_{datetime.now().strftime("%Y%m%d")}.xlsx"'

    return response


def export_inventory_to_pdf(inventory_qs):
    """
    Экспорт списка инвентаря в PDF.
    """
    font_name = _get_pdf_cyrillic_font()
    font_bold = 'PdfCyrillicBold' if font_name == 'PdfCyrillic' else font_name

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_bold,
        fontSize=24,
        textColor=colors.HexColor('#2C5F7C'),
        spaceAfter=30,
        alignment=1
    )
    normal_style = ParagraphStyle('PdfNormal', parent=styles['Normal'], fontName=font_name, fontSize=10)

    title = Paragraph("Каталог инвентаря СпортРент", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))

    date_text = "Дата: %s" % datetime.now().strftime('%d.%m.%Y %H:%M')
    elements.append(Paragraph(date_text, normal_style))
    elements.append(Spacer(1, 0.3 * inch))

    data = [['Название', 'Категория', 'Цена/день', 'Статус', 'Рейтинг']]
    for item in inventory_qs[:50]:
        data.append([
            (item.name[:30] if len(item.name) > 30 else item.name),
            item.category.name,
            "%s руб." % item.price_per_day,
            item.get_status_display(),
            "%.1f" % item.avg_rating if item.avg_rating else '-'
        ])

    table = Table(data, colWidths=[2.5 * inch, 1.5 * inch, 1 * inch, 1 * inch, 0.8 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C5F7C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), font_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F7F9')]),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))

    total_text = "Всего предметов: %d" % inventory_qs.count()
    elements.append(Paragraph(total_text, normal_style))

    # Генерация PDF
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inventory_{datetime.now().strftime("%Y%m%d")}.pdf"'

    return response


def export_stats_to_pdf(stats_data):
    """
    Экспорт статистики в PDF (для админки).
    """
    font_name = _get_pdf_cyrillic_font()
    font_bold = 'PdfCyrillicBold' if font_name == 'PdfCyrillic' else font_name

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_bold,
        fontSize=24,
        textColor=colors.HexColor('#2C5F7C'),
        spaceAfter=30,
        alignment=1
    )
    normal_style = ParagraphStyle('PdfNormal', parent=styles['Normal'], fontName=font_name, fontSize=10)

    title = Paragraph("Отчет СпортРент", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))

    period_text = "Период: %s" % datetime.now().strftime('%d.%m.%Y')
    elements.append(Paragraph(period_text, normal_style))
    elements.append(Spacer(1, 0.3 * inch))

    stats_table_data = [
        ['Показатель', 'Значение'],
        ['Всего пользователей', str(stats_data.get('total_users', 0))],
        ['Клиентов', str(stats_data.get('total_clients', 0))],
        ['Владельцев', str(stats_data.get('total_owners', 0))],
        ['Инвентаря', str(stats_data.get('total_inventory', 0))],
        ['Доступно', str(stats_data.get('available_inventory', 0))],
        ['Всего аренд', str(stats_data.get('total_rentals', 0))],
        ['Активных', str(stats_data.get('active_rentals', 0))],
        ['Завершенных', str(stats_data.get('completed_rentals', 0))],
        ['Общий доход', "%s руб." % stats_data.get('total_revenue', 0)],
    ]

    stats_table = Table(stats_table_data, colWidths=[3 * inch, 2 * inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C5F7C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), font_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F7F9')]),
    ]))

    elements.append(stats_table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="stats_{datetime.now().strftime("%Y%m%d")}.pdf"'

    return response